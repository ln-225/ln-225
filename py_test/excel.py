#!/usr/bin/env python3
# coding=utf-8

from glob import glob
import os
import openpyxl
from collections import namedtuple
import random
import time
import gc

info_template = './files/cmcc_tmp.xlsx'
input_dir = os.path.normpath('./files/cmcc')
output_dir = os.path.normpath('./files/cmcc_output')
if not os.path.exists(output_dir):
    os.mkdir(output_dir)

corp_info = list()
excel_tmp = openpyxl.load_workbook(info_template)
tmp_sheet1 = excel_tmp['Sheet1']
for row in tmp_sheet1.iter_rows(min_row=3, max_row=tmp_sheet1.max_row, min_col=1, max_col=tmp_sheet1.max_column):
    corp_info.append([row[i].value for i in range(0, tmp_sheet1.max_column)])

# for row in tmp_sheet1.rows:
#     corp_info.append([row[i].value for i in range(0, tmp_sheet1.max_column)])
#     #ws.append([row[i].value for i in range(0, tmp_sheet1.max_column)])
    
excel_tmp.close()

print('len[{}]'.format(len(corp_info)))

# exclude_set = {1, 2, 3, 4, 5, 6, 7, 26}
exclude_set = [0, 1, 2, 3, 4, 5, 6, 25]
for f in glob('{}/*.xlsx'.format(input_dir)):
    # if f.endswith('106550240018.xlsx'):
    outfile = os.path.join(output_dir, os.path.basename(f))
    if os.path.exists(outfile):
        os.remove(outfile)
    wb_out = openpyxl.Workbook(write_only=True)
    ws_out = wb_out.create_sheet("报送模板")
    b = time.time()
    print(f)
    cnt = 1
    wb = openpyxl.load_workbook(f, read_only=True)
    ws = wb.worksheets[0]
    # ws = wb['报送模板']
    print(ws.rows)
    for row in ws.rows:
        if cnt <=6:
            cnt +=1
            # continue
            n_row = list()
            for cell in row:
                n_row.append(cell.value)
            ws_out.append(n_row)
            # ws_out.append([row[i].value for i in range(0, ws.max_column)]) 
        else:
            if row[1].value is None:
                print('空数据，处理结束')
                break
            n_row = random.choice(corp_info)
            for i in exclude_set:
                n_row[i] = row[i].value
            ws_out.append(n_row)
        cnt +=1 
        if cnt % 100000 == 0:
            print('{} rows'.format(cnt))
    
    wb.close()
    wb_out.save(filename=outfile)
    wb_out.close()
    # 手动释放内存
    del wb, wb_out
    gc.collect()
    e = time.time()
    print('eclipse {}s'.format(e-b))



